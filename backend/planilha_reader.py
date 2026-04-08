"""
planilha_reader.py

Estratégia de leitura (em ordem de prioridade):

1. xlwings  — só usa se o Excel JÁ está com o arquivo aberto.
             NUNCA abre o arquivo em background.
2. openpyxl — lê o arquivo salvo em disco (funciona com Excel fechado).
             Requer que o OneDrive tenha sincronizado a versão mais recente.

Detecção de mudança: usa os.path.getmtime() (timestamp do arquivo).
Não abre o arquivo para calcular hash — resolve o Permission denied.

Estrutura de dados retornada:
  - dados_municipios: { cod_ibge: { status, tipo, municipio, regiao, os[], tecnicos[], resumo_status{} } }
  - dados_regionais:  { regiao:   { os[], tecnicos[], resumo_status{}, municipios[] } }
"""

import json
import os
import logging
from datetime import datetime
from collections import defaultdict

log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Detecção de mudança via mtime (sem abrir o arquivo)
# ─────────────────────────────────────────────────────────────────────────────

def _mtime(path: str) -> float:
    """Retorna timestamp de modificação do arquivo. Não abre o arquivo."""
    return os.path.getmtime(path)


# ─────────────────────────────────────────────────────────────────────────────
# Leitura via xlwings (somente se arquivo JA esta aberto no Excel)
# ─────────────────────────────────────────────────────────────────────────────

def _xlwings_arquivo_aberto(path: str):
    """
    Procura o arquivo entre os workbooks abertos no Excel.
    Retorna o workbook se encontrado, None caso contrario.
    Nunca abre um arquivo novo.
    """
    try:
        import xlwings as xw
        nome = os.path.basename(path).lower()
        path_norm = os.path.normpath(path).lower()

        for app in xw.apps:
            for wb in app.books:
                try:
                    if (wb.name.lower() == nome or
                            os.path.normpath(wb.fullname).lower() == path_norm):
                        return wb
                except Exception:
                    continue
    except ImportError:
        pass
    return None


def ler_via_xlwings(path, aba, col_ibge, col_status, col_tipo, col_municipio=None, col_regiao=None, col_tecnico=None, col_os=None):
    wb = _xlwings_arquivo_aberto(path)
    if wb is None:
        raise RuntimeError(f"Arquivo não encontrado aberto no Excel: {os.path.basename(path)}")
    log.info(f"[xlwings] Lendo arquivo já aberto: {wb.name}")
    try:
        sheet = wb.sheets[aba]
    except Exception:
        abas = [s.name for s in wb.sheets]
        raise ValueError(f"Aba '{aba}' não encontrada. Abas disponíveis: {', '.join(abas)}")
    dados_raw = sheet.used_range.value
    if not dados_raw or len(dados_raw) < 2:
        raise ValueError("Aba vazia ou sem linhas de dados")
    cabecalho = [str(c).strip() if c else "" for c in dados_raw[0]]
    try:
        idx_ibge   = cabecalho.index(col_ibge)
        idx_status = cabecalho.index(col_status)
    except ValueError as e:
        raise ValueError(f"Coluna não encontrada: {e}. Colunas na aba: {', '.join(cabecalho)}")
    idx_tipo      = cabecalho.index(col_tipo)      if col_tipo      in cabecalho else None
    idx_municipio = cabecalho.index(col_municipio) if col_municipio in cabecalho else None
    idx_regiao    = cabecalho.index(col_regiao)    if col_regiao    in cabecalho else None
    idx_tecnico   = cabecalho.index(col_tecnico)   if col_tecnico   in cabecalho else None
    idx_os        = cabecalho.index(col_os)        if col_os        in cabecalho else None
    resultado = _processar_linhas(dados_raw[1:], idx_ibge, idx_status, idx_tipo, idx_municipio, idx_regiao, idx_tecnico, idx_os)
    log.info(f"[xlwings] {len(resultado)} municípios lidos")
    return resultado

# ─────────────────────────────────────────────────────────────────────────────
# Leitura via openpyxl (arquivo em disco, Excel pode estar fechado)
# ─────────────────────────────────────────────────────────────────────────────

def ler_via_openpyxl(path, aba, col_ibge, col_status, col_tipo, col_municipio=None, col_regiao=None, col_tecnico=None, col_os=None):
    from openpyxl import load_workbook
    log.info(f"[openpyxl] Lendo arquivo em disco: {os.path.basename(path)}")
    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    except PermissionError:
        raise PermissionError(
            f"Arquivo bloqueado: {os.path.basename(path)}\n"
            "Salve a planilha no Excel (Ctrl+S) e clique em Recarregar."
        )
    except Exception as e:
        raise RuntimeError(f"Erro ao abrir planilha: {e}")
    if aba not in wb.sheetnames:
        abas = ", ".join(wb.sheetnames)
        wb.close()
        raise ValueError(f"Aba '{aba}' não encontrada. Abas disponíveis: {abas}")
    sheet = wb[aba]
    linhas = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not linhas:
        raise ValueError("Aba vazia")
    cabecalho = [str(c).strip() if c else "" for c in linhas[0]]
    try:
        idx_ibge   = cabecalho.index(col_ibge)
        idx_status = cabecalho.index(col_status)
    except ValueError as e:
        raise ValueError(f"Coluna não encontrada: {e}. Colunas na aba: {', '.join(cabecalho)}")
    idx_tipo      = cabecalho.index(col_tipo)      if col_tipo      in cabecalho else None
    idx_municipio = cabecalho.index(col_municipio) if col_municipio in cabecalho else None
    idx_regiao    = cabecalho.index(col_regiao)    if col_regiao    in cabecalho else None
    idx_tecnico   = cabecalho.index(col_tecnico)   if col_tecnico   in cabecalho else None
    idx_os        = cabecalho.index(col_os)        if col_os        in cabecalho else None
    resultado = _processar_linhas(linhas[1:], idx_ibge, idx_status, idx_tipo, idx_municipio, idx_regiao, idx_tecnico, idx_os)
    log.info(f"[openpyxl] {len(resultado)} municípios lidos")
    return resultado

# ─────────────────────────────────────────────────────────────────────────────
# Processamento comum de linhas
# ─────────────────────────────────────────────────────────────────────────────

def _processar_linhas(linhas, idx_ibge, idx_status, idx_tipo, idx_municipio=None, idx_regiao=None, idx_tecnico=None, idx_os=None):
    """
    Processa linhas da planilha e retorna dicionário de municípios.
    
    Retorna:
      { cod_ibge: { status, tipo, municipio, regiao, os[], tecnicos[], resumo_status{} } }
    """
    resultado = {}
    for linha in linhas:
        if not linha or linha[idx_ibge] is None:
            continue
        codigo = str(linha[idx_ibge]).strip()
        if codigo.endswith(".0"):
            codigo = codigo[:-2]
        codigo = codigo.zfill(7)
        if not codigo or codigo == "0000000":
            continue
        
        status = str(linha[idx_status]).strip() if (idx_status is not None and linha[idx_status]) else ""
        tipo = str(linha[idx_tipo]).strip() if (idx_tipo is not None and linha[idx_tipo]) else ""
        municipio = str(linha[idx_municipio]).strip() if (idx_municipio is not None and linha[idx_municipio]) else ""
        regiao = str(linha[idx_regiao]).strip() if (idx_regiao is not None and linha[idx_regiao]) else ""
        tecnico = str(linha[idx_tecnico]).strip() if (idx_tecnico is not None and linha[idx_tecnico]) else ""
        os_num = str(linha[idx_os]).strip() if (idx_os is not None and linha[idx_os]) else ""
        
        # Se já existe este município, agrega OS e técnicos
        if codigo in resultado:
            entry = resultado[codigo]
            if os_num and os_num not in entry["os"]:
                entry["os"].append(os_num)
            if tecnico and tecnico not in entry["tecnicos"]:
                entry["tecnicos"].append(tecnico)
            # Atualiza resumo de status
            if status:
                entry["resumo_status"][status] = entry["resumo_status"].get(status, 0) + 1
            # Atualiza status principal (último lido)
            entry["status"] = status if status else entry["status"]
            entry["tipo"] = tipo if tipo else entry["tipo"]
            entry["municipio"] = municipio if municipio else entry["municipio"]
            entry["regiao"] = regiao if regiao else entry["regiao"]
        else:
            resumo = {}
            if status:
                resumo[status] = 1
            resultado[codigo] = {
                "status": status,
                "tipo": tipo,
                "municipio": municipio,
                "regiao": regiao,
                "os": [os_num] if os_num else [],
                "tecnicos": [tecnico] if tecnico else [],
                "resumo_status": resumo,
            }
    return resultado

# ─────────────────────────────────────────────────────────────────────────────
# Funcao principal chamada pelo servidor
# ─────────────────────────────────────────────────────────────────────────────

def carregar_dados(config, forcar: bool = False):
    """
    Carrega dados com cache baseado em mtime.

    Prioridade de leitura:
    1. xlwings (arquivo aberto no Excel) - dados em memoria, sempre frescos
    2. openpyxl (arquivo em disco)       - dados da ultima versao salva

    Retorna: (dados_municipios, dados_regionais, fonte_str)
      - dados_municipios: { cod_ibge: { status, tipo, municipio, regiao, os[], tecnicos[], resumo_status{} } }
      - dados_regionais:  { regiao:   { os[], tecnicos[], resumo_status{}, municipios[], total_os, total_tecnicos } }
    """
    path          = config.PLANILHA_PATH
    aba           = config.PLANILHA_ABA
    col_ibge      = config.COLUNA_CODIGO_IBGE
    col_status    = config.COLUNA_STATUS
    cache_path    = config.CACHE_PATH
    col_tipo      = config.COLUNA_TIPO
    col_municipio = getattr(config, "COLUNA_MUNICIPIO", None)
    col_regiao    = getattr(config, "COLUNA_REGIAO", None)
    col_tecnico   = getattr(config, "COLUNA_TECNICO", None)
    col_os        = getattr(config, "COLUNA_OS", None)

    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Planilha nao encontrada: {path}\n"
            f"Verifique PLANILHA_PATH em config.py"
        )

    mtime_atual = _mtime(path)

    # Tenta cache (se nao forcado)
    if not forcar and os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cache = json.load(f)

            if cache.get("mtime") == mtime_atual:
                log.info(
                    f"[cache] Hit — {cache.get('total_municipios', '?')} municipios "
                    f"({cache.get('gerado_em', '?')})"
                )
                return cache["dados_municipios"], cache.get("dados_regionais", {}), cache.get("fonte", "cache")
            else:
                log.info("[cache] Arquivo modificado — reprocessando")
        except Exception as e:
            log.warning(f"[cache] Erro ao ler cache: {e}")

    # Le planilha
    dados_municipios  = None
    fonte  = None
    erros  = []

    # Tentativa 1: xlwings (apenas se arquivo estiver aberto no Excel)
    if _xlwings_arquivo_aberto(path) is not None:
        try:
            dados_municipios = ler_via_xlwings(path, aba, col_ibge, col_status, col_tipo, col_municipio, col_regiao, col_tecnico, col_os)
            fonte = "xlwings"
        except Exception as e:
            erros.append(f"xlwings: {e}")
            log.warning(f"[xlwings] Falhou: {e}")

    # Tentativa 2: openpyxl
    if dados_municipios is None:
        if erros:
            log.info("[openpyxl] Tentando fallback...")
        else:
            log.info("[openpyxl] Arquivo nao aberto no Excel — lendo disco")
        try:
            dados_municipios = ler_via_openpyxl(path, aba, col_ibge, col_status, col_tipo, col_municipio, col_regiao, col_tecnico, col_os)
            fonte = "openpyxl"
        except Exception as e:
            erros.append(f"openpyxl: {e}")
            log.error(f"[openpyxl] Falhou: {e}")

    if dados_municipios is None:
        raise RuntimeError(
            "Nao foi possivel ler a planilha.\n" + "\n".join(erros)
        )

    # Agrega dados por regional
    dados_regionais = _agregar_por_regional(dados_municipios)

    # Salva cache
    os.makedirs(os.path.dirname(os.path.abspath(cache_path)), exist_ok=True)
    cache_novo = {
        "mtime":             mtime_atual,
        "gerado_em":         datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "fonte":             fonte,
        "total_municipios":  len(dados_municipios),
        "dados_municipios":  dados_municipios,
        "dados_regionais":   dados_regionais,
    }
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_novo, f, ensure_ascii=False, indent=2)
        log.info(f"[cache] Salvo — {len(dados_municipios)} municipios via {fonte}")
    except Exception as e:
        log.warning(f"[cache] Nao foi possivel salvar cache: {e}")

    return dados_municipios, dados_regionais, fonte


def _agregar_por_regional(dados_municipios):
    """
    Agrega dados de municípios por regional.
    
    Retorna:
      { regiao: { os[], tecnicos[], resumo_status{}, municipios[], total_os, total_tecnicos } }
    """
    regionais = defaultdict(lambda: {
        "os": set(),
        "tecnicos": set(),
        "resumo_status": defaultdict(int),
        "municipios": [],
    })
    
    for cod_ibge, dados in dados_municipios.items():
        regiao = dados.get("regiao", "")
        if not regiao:
            continue
        
        r = regionais[regiao]
        r["municipios"].append(cod_ibge)
        
        # Agrega OS únicas
        for os_num in dados.get("os", []):
            r["os"].add(os_num)
        
        # Agrega técnicos únicos
        for tec in dados.get("tecnicos", []):
            r["tecnicos"].add(tec)
        
        # Agrega resumo de status
        for status, count in dados.get("resumo_status", {}).items():
            r["resumo_status"][status] += count
    
    # Converte sets para listas e adiciona totais
    resultado = {}
    for regiao, dados in regionais.items():
        resultado[regiao] = {
            "os": sorted(list(dados["os"])),
            "tecnicos": sorted(list(dados["tecnicos"])),
            "resumo_status": dict(dados["resumo_status"]),
            "municipios": dados["municipios"],
            "total_os": len(dados["os"]),
            "total_tecnicos": len(dados["tecnicos"]),
        }
    
    return resultado
