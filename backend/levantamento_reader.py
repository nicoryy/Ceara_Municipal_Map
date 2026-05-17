"""
levantamento_reader.py

Le a planilha de levantamento de um municipio especifico, formato:
    <LEVANTAMENTOS_BASE_PATH>/<NOME>/AUDITORIA/FECHAMENTO CENSO IP*.xlsm

Estrategia:
- Apenas openpyxl em modo somente-leitura (sem xlwings).
- Toda comparacao de nomes (pasta, arquivo, aba, colunas) e case-insensitive.
- Cache por municipio invalidado por mtime do .xlsm.
"""

import glob
import json
import os
import re
import logging
import warnings
import unicodedata
from datetime import datetime

log = logging.getLogger(__name__)


# -----------------------------------------------------------------------------
# Resolucao case-insensitive de pasta / arquivo / aba / colunas
# -----------------------------------------------------------------------------

def _normalize(nome: str) -> str:
    """Lowercase + remove acentos + mantem so [a-z0-9]. Para casamento robusto."""
    if not nome:
        return ""
    s = unicodedata.normalize("NFD", str(nome))
    s = s.encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", "", s.lower())


def _find_municipio_folder(base: str, nome: str) -> str:
    log.info(f"[busca] Procurando pasta do municipio '{nome}' em: {base}")
    if not os.path.isdir(base):
        log.warning(f"[busca] FALHA: diretorio base nao existe: {base}")
        raise FileNotFoundError(f"Diretorio base nao encontrado: {base}")
    alvo = _normalize(nome)
    for entry in os.listdir(base):
        if _normalize(entry) == alvo:
            full = os.path.join(base, entry)
            if os.path.isdir(full):
                log.info(f"[busca] OK pasta do municipio: {full}")
                return full
    log.warning(
        f"[busca] FALHA: pasta '{nome}' (normalizada='{alvo}') nao encontrada em {base}. "
        f"Entradas existentes: {os.listdir(base)}"
    )
    raise FileNotFoundError(f"Pasta do municipio nao encontrada: {nome}")


def _find_xlsm(folder: str) -> str:
    auditoria_alvo = os.path.join(folder, "AUDITORIA")
    log.info(f"[busca] Procurando subpasta AUDITORIA em: {folder}")
    auditoria = None
    for entry in os.listdir(folder):
        if entry.upper() == "AUDITORIA" and os.path.isdir(os.path.join(folder, entry)):
            auditoria = os.path.join(folder, entry)
            log.info(f"[busca] OK subpasta AUDITORIA: {auditoria}")
            break
    if auditoria is None:
        log.warning(
            f"[busca] FALHA: subpasta AUDITORIA nao encontrada em {folder}. "
            f"Subpastas existentes: {[e for e in os.listdir(folder) if os.path.isdir(os.path.join(folder, e))]}"
        )
        raise FileNotFoundError(f"Pasta AUDITORIA nao encontrada em: {folder}")

    log.info(f"[busca] Procurando arquivo 'FECHAMENTO CENSO IP*.xlsm' em: {auditoria}")
    candidatos = []
    for entry in os.listdir(auditoria):
        if entry.lower().endswith(".xlsm") and entry.upper().startswith("FECHAMENTO CENSO IP"):
            candidatos.append(os.path.join(auditoria, entry))
    if not candidatos:
        log.warning(
            f"[busca] FALHA: nenhum 'FECHAMENTO CENSO IP*.xlsm' em {auditoria}. "
            f"Arquivos existentes: {os.listdir(auditoria)}"
        )
        raise FileNotFoundError(
            f"Nenhum arquivo 'FECHAMENTO CENSO IP*.xlsm' encontrado em: {auditoria}"
        )
    candidatos.sort(key=os.path.getmtime, reverse=True)
    escolhido = candidatos[0]
    if len(candidatos) > 1:
        log.info(
            f"[busca] OK arquivo .xlsm (mais recente entre {len(candidatos)}): {escolhido}"
        )
    else:
        log.info(f"[busca] OK arquivo .xlsm: {escolhido}")
    return escolhido


def _find_sheet(wb, alvo: str) -> str:
    log.info(f"[busca] Procurando aba '{alvo}' (case-insensitive)")
    alvo_up = alvo.strip().upper()
    for nome in wb.sheetnames:
        if nome.strip().upper() == alvo_up:
            log.info(f"[busca] OK aba: '{nome}'")
            return nome
    log.warning(
        f"[busca] FALHA: aba '{alvo}' nao encontrada. "
        f"Abas disponiveis: {wb.sheetnames}"
    )
    raise ValueError(
        f"Aba '{alvo}' nao encontrada. Abas disponiveis: {', '.join(wb.sheetnames)}"
    )


def _indices_case_insensitive(cabecalho, nomes):
    """
    Retorna { nome_canonico_como_em_'nomes': idx_ou_None }.
    """
    norm = [str(c).strip().upper() if c is not None else "" for c in cabecalho]
    resultado = {}
    for nome in nomes:
        n = nome.strip().upper()
        resultado[nome] = norm.index(n) if n in norm else None
    return resultado


# -----------------------------------------------------------------------------
# Parsing de coordenadas (LATITUDE / LONGITUDE com virgula)
# -----------------------------------------------------------------------------

_MINUS_VARIANTS = ("−", "–", "—")  # −, –, —


def _parse_coord(v):
    """Aceita float, int ou string com ',' ou '.'. Retorna float ou None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f == f else None  # rejeita NaN
    s = str(v).strip()
    if not s:
        return None
    for m in _MINUS_VARIANTS:
        s = s.replace(m, "-")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _coord_valida(lat, lng):
    if lat is None or lng is None:
        return False
    if not (-90.0 <= lat <= 90.0):
        return False
    if not (-180.0 <= lng <= 180.0):
        return False
    if lat == 0.0 and lng == 0.0:
        return False
    return True


# -----------------------------------------------------------------------------
# Processamento das linhas
# -----------------------------------------------------------------------------

def _valor(linha, idx):
    if idx is None or idx >= len(linha):
        return ""
    v = linha[idx]
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y %H:%M:%S")
    return str(v).strip()


def _processar_pontos(linhas, idx_map, idx_lat, idx_lon, colunas):
    """
    linhas    -> iteravel de tuplas (linhas de dados, sem cabecalho)
    idx_map   -> { nome_canonico: idx_ou_None } das colunas a expor
    idx_lat   -> indice da coluna LATITUDE
    idx_lon   -> indice da coluna LONGITUDE
    colunas   -> lista ordenada de nomes canonicos (preserva ordem no dict)
    """
    pontos = []
    for linha in linhas:
        if not linha:
            continue
        lat = _parse_coord(linha[idx_lat]) if idx_lat is not None and idx_lat < len(linha) else None
        lng = _parse_coord(linha[idx_lon]) if idx_lon is not None and idx_lon < len(linha) else None
        if not _coord_valida(lat, lng):
            continue

        ponto = {"lat": lat, "lng": lng}
        for nome in colunas:
            ponto[nome] = _valor(linha, idx_map.get(nome))
        pontos.append(ponto)

    # Deteccao de pontos coincidentes (mesma coordenada arredondada a 6 casas)
    grupos = {}
    for p in pontos:
        chave = f"{round(p['lat'], 6)},{round(p['lng'], 6)}"
        grupos.setdefault(chave, []).append(p)
    for chave, grupo in grupos.items():
        stacked = len(grupo) > 1
        for p in grupo:
            p["__stacked"] = stacked
            p["__stack_size"] = len(grupo)
            p["__stack_key"] = chave

    return pontos


# -----------------------------------------------------------------------------
# Cache slug
# -----------------------------------------------------------------------------

def _slug(nome: str) -> str:
    return re.sub(r"[^\w\-]+", "_", nome.upper()).strip("_") or "MUNICIPIO"


# -----------------------------------------------------------------------------
# Funcao principal
# -----------------------------------------------------------------------------

def carregar_levantamento(config, nome_municipio: str):
    """
    Le o levantamento do municipio. Usa cache por mtime do .xlsm.
    Retorna lista de dicts (pontos).
    """
    base       = config.LEVANTAMENTOS_BASE_PATH
    cache_dir  = config.LEVANTAMENTOS_CACHE_DIR
    aba_alvo   = config.LEVANTAMENTO_ABA
    col_lat    = config.COLUNA_LAT
    col_lon    = config.COLUNA_LON
    colunas    = list(config.COLUNAS_LEVANTAMENTO)

    folder = _find_municipio_folder(base, nome_municipio)
    xlsm   = _find_xlsm(folder)
    mtime  = os.path.getmtime(xlsm)

    slug       = _slug(nome_municipio)
    cache_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), cache_dir, f"{slug}.json")

    # Tenta cache
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cache = json.load(f)
            if cache.get("mtime") == mtime and cache.get("arquivo") == xlsm:
                log.info(
                    f"[cache-lev] Hit {slug} - {cache.get('total_pontos','?')} pontos "
                    f"({cache.get('gerado_em','?')})"
                )
                return cache["dados"]
            log.info(f"[cache-lev] Invalidado {slug} (mtime mudou)")
        except Exception as e:
            log.warning(f"[cache-lev] Erro lendo cache {slug}: {e}")

    # Le planilha. Suprime UserWarnings do openpyxl sobre celulas marcadas como
    # data com serial fora do range (sao geradas durante iter_rows no modo read-only).
    from openpyxl import load_workbook
    log.info(f"[levantamento] Lendo {os.path.basename(xlsm)}")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        try:
            wb = load_workbook(xlsm, read_only=True, data_only=True, keep_vba=False)
        except PermissionError:
            raise PermissionError(
                f"Arquivo bloqueado: {os.path.basename(xlsm)}\n"
                "Feche o Excel ou aguarde o OneDrive sincronizar."
            )
        except Exception as e:
            raise RuntimeError(f"Erro ao abrir planilha: {e}")

        try:
            sheet_name = _find_sheet(wb, aba_alvo)
            sheet      = wb[sheet_name]
            linhas     = list(sheet.iter_rows(values_only=True))
        finally:
            wb.close()

    if not linhas:
        raise ValueError("Aba vazia")
    cabecalho = linhas[0]

    idx_map = _indices_case_insensitive(cabecalho, colunas)
    lat_map = _indices_case_insensitive(cabecalho, [col_lat])
    lon_map = _indices_case_insensitive(cabecalho, [col_lon])
    idx_lat = lat_map[col_lat]
    idx_lon = lon_map[col_lon]
    if idx_lat is None or idx_lon is None:
        raise ValueError(
            f"Colunas {col_lat}/{col_lon} nao encontradas. "
            f"Cabecalho: {', '.join(str(c) for c in cabecalho)}"
        )

    pontos = _processar_pontos(linhas[1:], idx_map, idx_lat, idx_lon, colunas)
    log.info(f"[levantamento] {len(pontos)} pontos validos em {slug}")

    # Salva cache
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    novo = {
        "mtime":         mtime,
        "arquivo":       xlsm,
        "municipio":     nome_municipio,
        "gerado_em":     datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "total_pontos":  len(pontos),
        "dados":         pontos,
    }
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(novo, f, ensure_ascii=False, indent=2)
        log.info(f"[cache-lev] Salvo {slug}")
    except Exception as e:
        log.warning(f"[cache-lev] Nao foi possivel salvar cache {slug}: {e}")

    return pontos
