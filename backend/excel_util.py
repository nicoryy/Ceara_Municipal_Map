"""
excel_util.py

Utilitarios compartilhados por planilha_reader.py e levantamento_reader.py
para lidar com planilhas .xlsm/.xlsx que podem estar abertas no Excel.

Problema que isso resolve:
- Quando o Excel abre um arquivo, ele cria na mesma pasta um arquivo-trava
  oculto "~$NOME.xlsm", com mtime novissimo e sem conteudo util. Se o
  buscador de arquivos nao filtrar esse nome, ele acaba escolhido como "o
  mais recente" e a leitura falha com PermissionError.
- Mesmo filtrando o "~$", o arquivo real pode estar com trava exclusiva
  dependendo de como o Excel/OneDrive o abriu. Nesses casos, copiar o
  arquivo para uma pasta temporaria costuma funcionar: o Excel normalmente
  mantem apenas uma trava de escrita (FILE_SHARE_READ), entao uma copia
  bruta ainda e permitida.
"""

import logging
import os
import shutil
import tempfile
import zipfile

log = logging.getLogger(__name__)


def eh_arquivo_trava_excel(nome: str) -> bool:
    """
    True para arquivos-trava/temporarios que nao devem ser tratados como
    planilhas de verdade: "~$NOME.xlsm" (trava do Excel) ou ".NOME.xlsx"
    (arquivo oculto, ex.: alguns sincronizadores).
    """
    base = os.path.basename(nome)
    return base.startswith("~$") or base.startswith(".")


def abrir_workbook_somente_leitura(caminho: str):
    """
    Abre `caminho` com openpyxl em modo somente-leitura. Se o arquivo
    estiver bloqueado (Excel/OneDrive segurando a trava), tenta copiar para
    uma pasta temporaria e ler a copia.

    Retorna (workbook, caminho_temp_ou_None). Quando caminho_temp nao e
    None, o chamador deve chama-lo de volta em limpar_temp() apos usar o
    workbook (idealmente em um finally, junto com wb.close()).

    Levanta PermissionError (com mensagem que nao pede pra fechar o Excel)
    se nem a leitura direta nem a copia funcionarem.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(caminho, read_only=True, data_only=True, keep_vba=False)
        return wb, None
    except (PermissionError, OSError, zipfile.BadZipFile) as e:
        log.info(f"[excel] Arquivo em uso ({e.__class__.__name__}) — lendo copia temporaria: {os.path.basename(caminho)}")

    tmp_dir = tempfile.mkdtemp(prefix="municipios_map_")
    tmp_path = os.path.join(tmp_dir, os.path.basename(caminho))
    try:
        shutil.copy2(caminho, tmp_path)
        wb = load_workbook(tmp_path, read_only=True, data_only=True, keep_vba=False)
        log.info(f"[excel] Copia temporaria lida com sucesso: {os.path.basename(caminho)}")
        return wb, tmp_dir
    except Exception as e:
        limpar_temp(tmp_dir)
        raise PermissionError(
            f"Planilha em uso: {os.path.basename(caminho)}. Tente novamente em instantes."
        ) from e


def limpar_temp(caminho_temp):
    """Remove o diretorio temporario criado por abrir_workbook_somente_leitura, se houver."""
    if not caminho_temp:
        return
    try:
        shutil.rmtree(caminho_temp, ignore_errors=True)
    except Exception as e:
        log.warning(f"[excel] Nao foi possivel limpar temp {caminho_temp}: {e}")
