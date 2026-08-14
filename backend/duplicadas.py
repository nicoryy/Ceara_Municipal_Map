"""
duplicadas.py

Deteccao de pontos duplicados (postes registrados mais de uma vez a poucos
metros de distancia) — porte puro-Python de detectar_pontos_proximos() do
script QGIS (C:\\Users\\Satel\\OneDrive - SATEL\\qgis_files\\Funcoes\\QGIS -
PORTAL.py, linhas 595-822), pra rodar em segundo plano no servidor sem
depender do runtime do QGIS nem de sklearn/pandas.

Diferencas conhecidas em relacao ao script original (ver plano da feature):
- A Regra A original tambem excluia "pontos ajustados para dentro" —
  artefato do pipeline geometrico do QGIS que nao existe aqui; aqui so
  excluimos os pontos fora do limite municipal (ver geometria_util).
- Projecao metrica local (equirretangular) em vez de UTM SIRGAS2000 —
  diferenca de ~0,1% no raio de 10m, so afeta pontos bem na borda do raio.
- CLUSTER_ID segue a ordem do menor indice de membro (deterministico), em
  vez da ordem arbitraria de rotulo do sklearn — muda a numeracao, nunca a
  composicao dos clusters.

Regras (nomeadas como no script original, A-G):
  A. exclui pontos fora do limite municipal
  B. coordenada com precisao total (identidade exata) preservada a parte
     da coordenada arredondada usada no clustering
  C. DBSCAN(eps=raio_metros, min_samples=2) == componentes conexos do grafo
     de adjacencia a <= raio_metros, descartando componentes de tamanho 1
     (prova: com min_samples=2 a relacao de vizinhanca e simetrica, entao
     todo ponto com >=1 vizinho e "core"; nao existe "border" isolado)
  D. dentro do cluster, colapsa coordenadas EXATAMENTE identicas ->
     representante = indice minimo
  E. cluster so e valido se sobrarem >=2 representantes apos D
  F. categoria por prioridade exclusiva: MEDICAO difere > DATA_REGISTRO
     difere (10 primeiros chars) > Nome_Cadastrador difere > "outros"
  G. rotulo sequencial CLUSTER_0001, CLUSTER_0002, ...
"""

import io
import logging
import math
import re
import unicodedata
from collections import defaultdict

from openpyxl import Workbook

log = logging.getLogger(__name__)

RAIO_METROS_PADRAO = 10.0

NOMES_ABAS = {
    "medicao":     "duplicadas_medicao",
    "data":        "duplicadas_data",
    "cadastrador": "duplicadas_cadastrador",
    "outros":      "duplicadas",
}
ORDEM_CATEGORIAS = ["medicao", "data", "cadastrador", "outros"]


# -----------------------------------------------------------------------------
# Regra A — exclusao por limite municipal
# -----------------------------------------------------------------------------

def _filtrar_dentro_do_limite(pontos, geometria):
    """Retorna (pontos_validos, total_excluidos). Se geometria for None
    (limite desconhecido), nao exclui ninguem."""
    if geometria is None:
        return list(pontos), 0
    from geometria_util import ponto_dentro_geometria
    validos = []
    excluidos = 0
    for p in pontos:
        if ponto_dentro_geometria(p["lat"], p["lng"], geometria):
            validos.append(p)
        else:
            excluidos += 1
    return validos, excluidos


# -----------------------------------------------------------------------------
# Projecao metrica local (Regra B usa lat/lng cheios; clustering usa isto)
# -----------------------------------------------------------------------------

def _projetar(pontos):
    lat0 = sum(p["lat"] for p in pontos) / len(pontos)
    lon0 = sum(p["lng"] for p in pontos) / len(pontos)
    cos_lat0 = math.cos(math.radians(lat0)) or 1e-9
    return [
        ((p["lng"] - lon0) * 111320.0 * cos_lat0, (p["lat"] - lat0) * 110540.0)
        for p in pontos
    ]


# -----------------------------------------------------------------------------
# Regra C — clustering por proximidade (grid + union-find == DBSCAN eps/min=2)
# -----------------------------------------------------------------------------

class _UnionFind:
    def __init__(self, n):
        self.pai = list(range(n))

    def find(self, i):
        while self.pai[i] != i:
            self.pai[i] = self.pai[self.pai[i]]
            i = self.pai[i]
        return i

    def unir(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.pai[ra] = rb


def _clusterizar(coords_metros, raio_metros):
    """coords_metros: lista de (x, y) em metros. Retorna lista de listas de
    indices — componentes conexos do grafo de adjacencia a <= raio_metros,
    ja descartando componentes de tamanho 1 (equivalente a DBSCAN com
    min_samples=2 — ver docstring do modulo)."""
    n = len(coords_metros)
    if n == 0:
        return []

    lado = raio_metros
    grade = defaultdict(list)
    for i, (x, y) in enumerate(coords_metros):
        grade[(int(x // lado), int(y // lado))].append(i)

    uf = _UnionFind(n)
    raio2 = raio_metros * raio_metros
    for (cx, cy), indices in grade.items():
        vizinhos = []
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                vizinhos.extend(grade.get((cx + dx, cy + dy), ()))
        for i in indices:
            xi, yi = coords_metros[i]
            for j in vizinhos:
                if j <= i:
                    continue
                xj, yj = coords_metros[j]
                dx, dy = xi - xj, yi - yj
                if dx * dx + dy * dy <= raio2:
                    uf.unir(i, j)

    grupos = defaultdict(list)
    for i in range(n):
        grupos[uf.find(i)].append(i)
    return [idxs for idxs in grupos.values() if len(idxs) > 1]


# -----------------------------------------------------------------------------
# Regra D/E — colapso de coordenadas identicas + filtro de validade
# -----------------------------------------------------------------------------

def _colapsar_identicos(indices, pontos):
    grupos = defaultdict(list)
    for i in indices:
        p = pontos[i]
        grupos[(p["lat"], p["lng"])].append(i)
    return sorted(min(g) for g in grupos.values())


# -----------------------------------------------------------------------------
# Regra F — categorizacao por prioridade exclusiva
# -----------------------------------------------------------------------------

def _categorizar(indices_repr, pontos):
    valores_medicao = {str(pontos[i].get("MEDICAO") or "").strip() for i in indices_repr}
    if len(valores_medicao) > 1:
        return "medicao"

    valores_data = {str(pontos[i].get("DATA_REGISTRO") or "").strip()[:10] for i in indices_repr}
    if len(valores_data) > 1:
        return "data"

    valores_cadastrador = {str(pontos[i].get("Nome_Cadastrador") or "").strip().upper() for i in indices_repr}
    if len(valores_cadastrador) > 1:
        return "cadastrador"

    return "outros"


# -----------------------------------------------------------------------------
# Orquestracao
# -----------------------------------------------------------------------------

def detectar_duplicadas(pontos, geometria=None, raio_metros=RAIO_METROS_PADRAO):
    """
    pontos -> lista de dicts como os que carregar_levantamento() devolve
    (precisam ter "lat"/"lng" e as colunas usadas em _categorizar).

    Retorna {
      "clusters": [{"cluster_id", "categoria", "pontos": [pontos...]}],
      "excluidos_fora_limite": int,
      "total_considerados": int,
    }
    """
    pontos_validos, excluidos = _filtrar_dentro_do_limite(pontos, geometria)
    pontos_validos = [
        p for p in pontos_validos
        if isinstance(p.get("lat"), (int, float)) and isinstance(p.get("lng"), (int, float))
    ]

    if len(pontos_validos) < 2:
        return {"clusters": [], "excluidos_fora_limite": excluidos, "total_considerados": len(pontos_validos)}

    coords_metros = _projetar(pontos_validos)
    componentes = _clusterizar(coords_metros, raio_metros)

    clusters_validos = []
    for indices in componentes:
        representantes = _colapsar_identicos(indices, pontos_validos)
        if len(representantes) < 2:
            continue
        categoria = _categorizar(representantes, pontos_validos)
        clusters_validos.append((representantes, categoria))

    # Regra G — numeracao deterministica pela ordem do menor indice de membro
    clusters_validos.sort(key=lambda t: min(t[0]))

    resultado = []
    for n, (representantes, categoria) in enumerate(clusters_validos, start=1):
        resultado.append({
            "cluster_id": f"CLUSTER_{n:04d}",
            "categoria":  categoria,
            "pontos":     [pontos_validos[i] for i in representantes],
        })

    log.info(
        f"[duplicadas] {len(resultado)} clusters validos "
        f"({sum(len(c['pontos']) for c in resultado)} pontos) — "
        f"{excluidos} pontos fora do limite excluidos"
    )
    return {
        "clusters": resultado,
        "excluidos_fora_limite": excluidos,
        "total_considerados": len(pontos_validos),
    }


# -----------------------------------------------------------------------------
# Exportacao XLSX — paridade com o script QGIS (ver QGIS - PORTAL.py:903-970)
# -----------------------------------------------------------------------------

def exportar_xlsx(resultado, colunas_levantamento):
    """
    Um workbook, ate 4 abas (so as que tiverem clusters), cada cluster
    aparece em exatamente uma aba — a aba "duplicadas" e a categoria
    "outros" (tudo igual), nao a uniao de todas. Linhas = representantes,
    ordenadas por CLUSTER_ID.
    """
    por_categoria = defaultdict(list)
    for cluster in resultado["clusters"]:
        for p in cluster["pontos"]:
            por_categoria[cluster["categoria"]].append((cluster["cluster_id"], p))

    wb = Workbook()
    wb.remove(wb.active)

    cabecalho = list(colunas_levantamento) + ["CLUSTER_ID", "X_Coordinate", "Y_Coordinate"]
    abas_escritas = 0
    for categoria in ORDEM_CATEGORIAS:
        linhas = por_categoria.get(categoria)
        if not linhas:
            continue
        linhas.sort(key=lambda t: t[0])
        ws = wb.create_sheet(NOMES_ABAS[categoria])
        ws.append(cabecalho)
        for cluster_id, p in linhas:
            linha = [p.get(col, "") for col in colunas_levantamento]
            linha += [cluster_id, p.get("lng", ""), p.get("lat", "")]
            ws.append(linha)
        abas_escritas += 1

    if abas_escritas == 0:
        ws = wb.create_sheet("duplicadas")
        ws.append(cabecalho)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# -----------------------------------------------------------------------------
# Nome do arquivo — mesma convencao do script QGIS (linhas 914-918)
# -----------------------------------------------------------------------------

_REGIONAL_RE = re.compile(r"^FORTALEZA\s*-\s*REGIONAL\s*(\d+)$", re.IGNORECASE)


def _remover_acentos(s: str) -> str:
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def nome_arquivo_duplicados(nome_municipio: str) -> str:
    m = _REGIONAL_RE.match(nome_municipio.strip())
    if m:
        return f"duplicados_Regional_{m.group(1)}.xlsx"
    limpo = _remover_acentos(nome_municipio).upper().replace(" ", "_")
    limpo = re.sub(r"[^A-Z0-9_]+", "", limpo) or "MUNICIPIO"
    return f"{limpo}_duplicados.xlsx"
